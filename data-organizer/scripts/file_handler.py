#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
文件处理模块
支持 PDF（文本+OCR）、Word、Excel、CSV、JSON、图片
使用 PaddleOCR 处理扫描件
"""

import os
import json
import csv
import numpy as np
from pathlib import Path
from datetime import datetime
from typing import Dict, Any

try:
    from pdf2image import convert_from_path
except ImportError:
    convert_from_path = None

try:
    from paddleocr import PaddleOCR
except ImportError:
    PaddleOCR = None

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

try:
    import docx
except ImportError:
    docx = None

try:
    from PIL import Image
except ImportError:
    Image = None


FILE_TYPE_MAP = {
    '.pdf': 'pdf', '.xlsx': 'excel', '.xls': 'excel', '.docx': 'word',
    '.csv': 'csv', '.json': 'json', '.txt': 'text', '.md': 'markdown',
    '.log': 'text', '.png': 'image', '.jpg': 'image', '.jpeg': 'image',
    '.gif': 'image', '.bmp': 'image',
}


def get_file_type(file_path: str) -> str:
    ext = Path(file_path).suffix.lower()
    return FILE_TYPE_MAP.get(ext, 'unknown')


def get_file_metadata(file_path: str) -> Dict[str, Any]:
    path = Path(file_path)
    stat = path.stat()
    return {
        'size': stat.st_size,
        'modified': datetime.fromtimestamp(stat.st_mtime).isoformat(),
        'extension': path.suffix.lower(),
    }


def extract_pdf_content(file_path: str) -> str:
    """提取 PDF 内容 - 使用 PaddleOCR（支持扫描件）"""
    if PaddleOCR is None:
        return "[需要安装 paddleocr: pip install paddleocr paddlepaddle]"

    if convert_from_path is None:
        return "[需要安装 pdf2image: pip install pdf2image]"

    try:
        print(f"[FILE_HANDLER] 初始化 PaddleOCR...")
        ocr = PaddleOCR(use_angle_cls=True, lang='ch', show_log=False)

        print(f"[FILE_HANDLER] PDF 转图片 (DPI=100)...")
        images = convert_from_path(file_path, dpi=100)
        print(f"[FILE_HANDLER] 共 {len(images)} 页，使用 PaddleOCR...")

        text_parts = []
        for page_idx, image in enumerate(images):
            print(f"[FILE_HANDLER] 处理页面 {page_idx+1}/{len(images)}...")
            img_array = np.array(image)
            result = ocr.ocr(img_array)

            if result and result[0]:
                lines = []
                for line in result[0]:
                    if line and len(line) >= 2:
                        text = line[1][0]
                        lines.append(text)
                if lines:
                    text_parts.append(f"[页面 {page_idx+1}]\n" + '\n'.join(lines))
                    print(f"[FILE_HANDLER] 页面 {page_idx+1}: 识别 {len(lines)} 行")

        if text_parts:
            print(f"[FILE_HANDLER] PaddleOCR 成功 {len(text_parts)}/{len(images)} 页")
            return '\n\n'.join(text_parts)
        return "[PDF 无法提取内容]"

    except Exception as e:
        import traceback
        traceback.print_exc()
        return f"[PDF 处理错误: {str(e)}]"


def extract_excel_content(file_path: str) -> str:
    """提取 Excel 内容"""
    if load_workbook is None:
        return "[需要安装 openpyxl: pip install openpyxl]"

    try:
        wb = load_workbook(file_path, read_only=True)
        content = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            content.append(f"[工作表: {sheet_name}]")
            for row in ws.iter_rows(values_only=True):
                content.append('\t'.join(str(c) if c else '' for c in row))
        wb.close()
        return '\n'.join(content)
    except Exception as e:
        return f"[Excel 错误: {e}]"


def extract_word_content(file_path: str) -> str:
    """提取 Word 内容"""
    if docx is None:
        return "[需要安装 python-docx: pip install python-docx]"

    try:
        doc = docx.Document(file_path)
        paras = [p.text.strip() for p in doc.paragraphs if p.text.strip()]
        return '\n'.join(paras)
    except Exception as e:
        return f"[Word 错误: {e}]"


def extract_csv_content(file_path: str) -> str:
    """提取 CSV 内容"""
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            return '\n'.join(','.join(r) for r in csv.reader(f))
    except Exception as e:
        return f"[CSV 错误: {e}]"


def extract_json_content(file_path: str) -> str:
    """提取 JSON 内容"""
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            return json.dumps(json.load(f), ensure_ascii=False, indent=2)
    except Exception as e:
        return f"[JSON 错误: {e}]"


def extract_image_info(file_path: str) -> Dict[str, Any]:
    """提取图片信息"""
    if Image is None:
        return {"error": "需要安装 pillow"}
    try:
        img = Image.open(file_path)
        return {"format": img.format, "size": img.size, "mode": img.mode}
    except Exception as e:
        return {"error": str(e)}


def extract_text_content(file_path: str) -> str:
    """提取文本文件内容"""
    for enc in ['utf-8', 'gbk', 'gb2312', 'latin-1']:
        try:
            with open(file_path, 'r', encoding=enc) as f:
                return f.read().strip()
        except UnicodeDecodeError:
            continue
    return "[文本解码错误]"


def process_file(file_path: str) -> Dict[str, Any]:
    """处理单个文件"""
    file_type = get_file_type(file_path)
    metadata = get_file_metadata(file_path)

    if file_type == 'pdf':
        content = extract_pdf_content(file_path)
    elif file_type == 'excel':
        content = extract_excel_content(file_path)
    elif file_type == 'word':
        content = extract_word_content(file_path)
    elif file_type == 'csv':
        content = extract_csv_content(file_path)
    elif file_type == 'json':
        content = extract_json_content(file_path)
    elif file_type == 'image':
        content = ""
        metadata['image_info'] = extract_image_info(file_path)
    else:
        content = extract_text_content(file_path)

    return {
        'path': str(Path(file_path).resolve()),
        'type': file_type,
        'name': Path(file_path).name,
        'content': content,
        'metadata': metadata,
    }
